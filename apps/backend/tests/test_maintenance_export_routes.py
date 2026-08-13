"""Ruta de exportación de Órdenes de Trabajo, de punta a punta.

Comprueba que el archivo descargado refleja el mismo recorte que el tablero:
si el filtro del endpoint no coincidiera con el de la pantalla, el usuario se
llevaría un listado distinto del que está viendo.
"""

from uuid import uuid4

import pytest
from src.features.machine.domain.entities import Machine, MachineStatus
from src.features.machine.infrastructure.repositories import MachineRepository
from src.features.maintenance.domain.entities import (
    FailureCategory,
    MaintenanceOrder,
    MaintenanceStatus,
)
from src.features.maintenance.infrastructure.repositories import (
    MaintenanceOrderRepository,
)
from src.features.maintenance.infrastructure.routes import (
    _load_orders_for_export,
    export_order_sheet,
    export_orders,
)
from src.features.user.domain.entities import UserMetadata, UserRole
from src.features.user.infrastructure.repositories import UserRepository


@pytest.fixture
async def taller(test_uow):
    """Dos máquinas, un mecánico y tres OT en distintos estados."""
    async with test_uow:
        mechanic = await UserRepository(test_uow).save(
            UserMetadata(
                better_auth_user_id="auth-mecanico",
                role=UserRole.MECANICO,
            )
        )

        machine_repo = MachineRepository(test_uow)
        excavadora = await machine_repo.save(
            Machine(
                code="EXC-001",
                motor_serial="MTR-001",
                brand="Caterpillar",
                model="320D",
                manufacture_year=2019,
                current_horometer=1250.5,
                status=MachineStatus.ACTIVA,
            )
        )
        camion = await machine_repo.save(
            Machine(
                code="CAM-014",
                motor_serial="MTR-014",
                brand="Volvo",
                model="FMX 480",
                manufacture_year=2021,
                current_horometer=86500.0,
                status=MachineStatus.ACTIVA,
            )
        )

        order_repo = MaintenanceOrderRepository(test_uow)
        for machine, status, category in [
            (excavadora, MaintenanceStatus.PROGRAMADO, FailureCategory.MOTOR),
            (excavadora, MaintenanceStatus.LIQUIDADO, FailureCategory.FRENOS),
            (camion, MaintenanceStatus.PROGRAMADO, FailureCategory.MOTOR),
        ]:
            await order_repo.save(
                MaintenanceOrder(
                    machine_id=machine.id,
                    description=f"Intervención en {machine.code}",
                    status=status,
                    assigned_mechanic_id=mechanic.id,
                    failure_category=category,
                )
            )

        await test_uow.commit()

    return {"uow": test_uow, "excavadora": excavadora, "camion": camion}


class TestLoadOrdersForExport:
    @pytest.mark.asyncio
    async def test_sin_filtros_carga_todas_las_ot(self, taller):
        orders, scope = await _load_orders_for_export(taller["uow"])

        assert len(orders) == 3
        assert scope == "Todas las OT"

    @pytest.mark.asyncio
    async def test_filtra_por_estado(self, taller):
        orders, scope = await _load_orders_for_export(
            taller["uow"], status_filter=MaintenanceStatus.PROGRAMADO.value
        )

        assert len(orders) == 2
        assert "Programado" in scope

    @pytest.mark.asyncio
    async def test_filtra_por_maquina(self, taller):
        orders, scope = await _load_orders_for_export(
            taller["uow"], machine_id=str(taller["excavadora"].id)
        )

        assert len(orders) == 2
        assert {o.machine.code for o in orders} == {"EXC-001"}
        assert "máquina" in scope

    @pytest.mark.asyncio
    async def test_filtra_por_clasificacion_de_falla(self, taller):
        orders, _ = await _load_orders_for_export(
            taller["uow"], failure_category=FailureCategory.MOTOR.value
        )

        assert len(orders) == 2

    @pytest.mark.asyncio
    async def test_combina_estado_y_maquina(self, taller):
        orders, _ = await _load_orders_for_export(
            taller["uow"],
            status_filter=MaintenanceStatus.LIQUIDADO.value,
            machine_id=str(taller["excavadora"].id),
        )

        assert len(orders) == 1

    @pytest.mark.asyncio
    async def test_el_centinela_all_no_filtra(self, taller):
        orders, scope = await _load_orders_for_export(
            taller["uow"], status_filter="ALL", machine_id="ALL"
        )

        assert len(orders) == 3
        assert scope == "Todas las OT"

    @pytest.mark.asyncio
    async def test_id_de_maquina_invalido_es_error_de_peticion(self, taller):
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as exc:
            await _load_orders_for_export(taller["uow"], machine_id="no-es-un-uuid")

        assert exc.value.status_code == 400

    @pytest.mark.asyncio
    async def test_orden_descendente_por_fecha(self, taller):
        orders, _ = await _load_orders_for_export(taller["uow"])

        fechas = [o.created_at for o in orders]
        assert fechas == sorted(fechas, reverse=True)


class TestExportOrdersRoute:
    @pytest.mark.asyncio
    @pytest.mark.parametrize(
        "fmt,media_type,magic",
        [
            ("xlsx", "spreadsheetml", b"PK"),
            ("pdf", "application/pdf", b"%PDF-"),
            ("csv", "text/csv", None),
        ],
    )
    async def test_devuelve_el_formato_pedido(self, taller, fmt, media_type, magic):
        response = await export_orders(format=fmt, uow=taller["uow"])

        assert media_type in response.media_type
        assert "attachment" in response.headers["content-disposition"]
        assert f".{fmt}" in response.headers["content-disposition"]
        if magic:
            assert response.body[:len(magic)] == magic

    @pytest.mark.asyncio
    async def test_formato_por_defecto_es_excel(self, taller):
        response = await export_orders(uow=taller["uow"])

        assert response.body[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_formato_invalido_es_error_de_peticion(self, taller):
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as exc:
            await export_orders(format="docx", uow=taller["uow"])

        assert exc.value.status_code == 400

    @pytest.mark.asyncio
    async def test_el_filtro_llega_al_archivo(self, taller):
        """El Excel filtrado por máquina no debe contener la otra máquina."""
        import io

        from openpyxl import load_workbook

        response = await export_orders(
            format="xlsx", machine_id=str(taller["excavadora"].id), uow=taller["uow"]
        )
        sheet = load_workbook(io.BytesIO(response.body)).active
        text = "\n".join(
            str(cell)
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )

        assert "EXC-001" in text
        assert "CAM-014" not in text


class TestExportOrderSheetRoute:
    @pytest.mark.asyncio
    async def test_descarga_la_hoja_de_una_ot(self, taller):
        orders, _ = await _load_orders_for_export(taller["uow"])
        target = orders[0]

        response = await export_order_sheet(str(target.id), uow=taller["uow"])

        assert response.media_type == "application/pdf"
        assert response.body[:5] == b"%PDF-"
        # El nombre del archivo lleva el folio de la OT.
        assert "ot-" in response.headers["content-disposition"]

    @pytest.mark.asyncio
    async def test_id_invalido_es_error_de_peticion(self, taller):
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as exc:
            await export_order_sheet("no-es-un-uuid", uow=taller["uow"])

        assert exc.value.status_code == 400

    @pytest.mark.asyncio
    async def test_ot_inexistente_propaga_el_no_encontrado(self, taller):
        from src.features.maintenance.domain.exceptions import (
            MaintenanceNotFoundException,
        )

        with pytest.raises(MaintenanceNotFoundException):
            await export_order_sheet(str(uuid4()), uow=taller["uow"])
