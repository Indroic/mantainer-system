"""Exportación de Órdenes de Trabajo en Excel, CSV y PDF (spec 4.4).

Cubre el caso delicado del costo: ``unit_cost_at_time`` solo se congela al
liquidar la OT, así que los importes de una OT abierta son una ESTIMACIÓN tomada
del catálogo vigente y el documento debe advertirlo.
"""

import io
from datetime import datetime, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from src.features.inventory.application.dtos import SparePartResponse
from src.features.machine.application.dtos import MachineResponse
from src.features.maintenance.application.dtos import (
    MaintenanceResponse,
    MaintenanceSparePartResponse,
)
from src.features.maintenance.domain.entities import FailureCategory, MaintenanceStatus
from src.features.maintenance.infrastructure.exporters import (
    ORDER_COLUMNS,
    order_code,
    order_parts_total,
    order_units_total,
    render_order_sheet_pdf,
    render_orders_csv,
    render_orders_pdf,
    render_orders_xlsx,
    status_label,
)

NOW = datetime.now(timezone.utc)


def _machine(code: str = "EXC-001") -> MachineResponse:
    return MachineResponse(
        id=uuid4(),
        code=code,
        motor_serial="MTR-9F82K1",
        brand="Caterpillar",
        model="320D",
        manufacture_year=2019,
        current_horometer=1250.5,
        status="ACTIVA",
        horometer_unit="Horas",
        description="Excavadora hidráulica",
        location="Patio Principal",
        machine_type_id=None,
        machine_type_name="Excavadora",
        created_at=NOW,
        updated_at=NOW,
        is_active=True,
    )


def _part(code: str, name: str, unit_cost: float) -> SparePartResponse:
    return SparePartResponse(
        id=uuid4(),
        code=code,
        name=name,
        stock_minimum=2,
        unit_cost=unit_cost,
        stock_current=10,
        created_at=NOW,
        updated_at=NOW,
        is_active=True,
    )


def _order(
    *,
    status: MaintenanceStatus,
    unit_cost_at_time: float | None,
    quantity: int = 2,
    catalog_cost: float = 12.25,
    work_performed: str | None = None,
    machine_code: str = "EXC-001",
) -> MaintenanceResponse:
    return MaintenanceResponse(
        id=uuid4(),
        machine_id=uuid4(),
        description="Cambio de aceite y filtros del motor",
        status=status,
        assigned_mechanic_id=uuid4(),
        assigned_mechanic_name="Juan Pérez",
        next_service_horometer=1500.0,
        failure_category=FailureCategory.MANTENIMIENTO_PREVENTIVO,
        failure_category_label="Mantenimiento Preventivo",
        work_performed=work_performed,
        created_by="auth-1",
        created_by_name="Ana Torres",
        spare_parts=[
            MaintenanceSparePartResponse(
                id=uuid4(),
                spare_part_id=uuid4(),
                quantity_requested=quantity,
                quantity_returned=0,
                quantity=quantity,
                unit_cost_at_time=unit_cost_at_time,
                spare_part=_part("FLT-001", "Filtro de aceite", catalog_cost),
            )
        ],
        machine=_machine(machine_code),
        solvencies=[],
        created_at=NOW,
        updated_at=NOW,
        is_active=True,
    )


@pytest.fixture
def liquidada() -> MaintenanceResponse:
    """OT cerrada: el costo histórico ya está congelado."""
    return _order(
        status=MaintenanceStatus.LIQUIDADO,
        unit_cost_at_time=45.50,
        quantity=2,
        work_performed="Se drenaron 20L de aceite y se sustituyeron 2 filtros.",
    )


@pytest.fixture
def abierta() -> MaintenanceResponse:
    """OT en ejecución: sin costo congelado, el importe es una estimación."""
    return _order(
        status=MaintenanceStatus.EN_EJECUCION,
        unit_cost_at_time=None,
        quantity=3,
        catalog_cost=12.25,
        machine_code="EXC-002",
    )


class TestOrderCosts:
    def test_ot_liquidada_usa_el_costo_historico(self, liquidada):
        assert order_parts_total(liquidada) == pytest.approx(91.0)  # 2 x 45,50

    def test_ot_abierta_estima_con_el_catalogo(self, abierta):
        """Sin costo congelado se estima con el catálogo, en vez de informar 0,00."""
        assert order_parts_total(abierta) == pytest.approx(36.75)  # 3 x 12,25

    def test_cuenta_las_unidades_solicitadas(self, liquidada, abierta):
        assert order_units_total(liquidada) == 2
        assert order_units_total(abierta) == 3

    def test_ot_sin_repuestos_no_falla(self):
        vacia = _order(status=MaintenanceStatus.PROGRAMADO, unit_cost_at_time=None)
        vacia.spare_parts = []

        assert order_parts_total(vacia) == 0
        assert order_units_total(vacia) == 0

    def test_folio_estable_y_legible(self, liquidada):
        code = order_code(liquidada)

        assert code.startswith("OT-")
        assert code == order_code(liquidada)  # determinista

    @pytest.mark.parametrize(
        "status,expected",
        [
            (MaintenanceStatus.PROGRAMADO, "Programado"),
            (MaintenanceStatus.EN_EJECUCION, "En ejecución"),
            (MaintenanceStatus.LIQUIDADO, "Liquidado"),
        ],
    )
    def test_etiqueta_de_estado_en_espanol(self, status, expected):
        assert status_label(status) == expected


class TestOrdersXlsx:
    def test_genera_un_libro_valido_y_reabrible(self, liquidada, abierta):
        content = render_orders_xlsx([liquidada, abierta])

        assert content[:2] == b"PK"  # un XLSX es un ZIP

        sheet = load_workbook(io.BytesIO(content)).active
        text = "\n".join(
            str(cell)
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        assert order_code(liquidada) in text
        assert order_code(abierta) in text
        assert "EXC-001" in text
        assert "Liquidado" in text

    def test_incluye_todas_las_columnas_declaradas(self, liquidada):
        sheet = load_workbook(io.BytesIO(render_orders_xlsx([liquidada]))).active

        headers = {
            str(cell)
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        }
        for column in ORDER_COLUMNS:
            assert column.header in headers

    def test_advierte_cuando_hay_importes_estimados(self, abierta):
        sheet = load_workbook(io.BytesIO(render_orders_xlsx([abierta]))).active

        text = "\n".join(
            str(cell)
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        assert "estimados" in text

    def test_totaliza_el_costo_de_todas_las_ot(self, liquidada, abierta):
        sheet = load_workbook(io.BytesIO(render_orders_xlsx([liquidada, abierta]))).active

        numeros = [
            cell
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if isinstance(cell, (int, float))
        ]
        # 91,00 + 36,75 debe aparecer como total.
        assert pytest.approx(127.75) in numeros

    def test_listado_vacio_no_revienta(self):
        content = render_orders_xlsx([])
        assert content[:2] == b"PK"


class TestOrdersCsv:
    def test_cabecera_y_una_fila_por_ot(self, liquidada, abierta):
        text = render_orders_csv([liquidada, abierta])
        lines = text.splitlines()

        assert len(lines) == 3  # cabecera + 2 OT
        assert ORDER_COLUMNS[0].header in lines[0]

    def test_lleva_bom_para_que_excel_respete_los_acentos(self, liquidada):
        assert render_orders_csv([liquidada]).startswith("﻿")

    def test_columnas_del_csv_coinciden_con_las_del_excel(self, liquidada):
        header = render_orders_csv([liquidada]).splitlines()[0].lstrip("﻿")

        assert header.split(",")[0] == ORDER_COLUMNS[0].header
        assert len(header.split(",")) == len(ORDER_COLUMNS)


class TestOrdersPdf:
    def test_genera_un_pdf_valido(self, liquidada, abierta):
        content = render_orders_pdf([liquidada, abierta])

        assert content[:5] == b"%PDF-"
        assert len(content) > 1000

    def test_listado_vacio_no_revienta(self):
        assert render_orders_pdf([])[:5] == b"%PDF-"


class TestOrderSheetPdf:
    def test_hoja_de_ot_liquidada(self, liquidada):
        content = render_order_sheet_pdf(liquidada)

        assert content[:5] == b"%PDF-"
        assert len(content) > 1000

    def test_hoja_de_ot_abierta_sin_trabajo_realizado(self, abierta):
        """Una OT sin liquidar no tiene 'trabajo realizado': no debe fallar."""
        assert abierta.work_performed is None

        content = render_order_sheet_pdf(abierta)
        assert content[:5] == b"%PDF-"

    def test_hoja_sin_maquina_asociada(self, liquidada):
        """`_to_maintenance_response` deja `machine` en None si no se pudo cargar."""
        liquidada.machine = None

        assert render_order_sheet_pdf(liquidada)[:5] == b"%PDF-"

    def test_hoja_sin_repuestos(self, liquidada):
        liquidada.spare_parts = []

        assert render_order_sheet_pdf(liquidada)[:5] == b"%PDF-"
