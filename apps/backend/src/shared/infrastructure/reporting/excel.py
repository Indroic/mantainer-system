"""Constructor de libros Excel (XLSX) profesionales del SGMM.

Resuelve los requisitos de la spec 4.4 en un solo lugar: encabezados claros,
auto-ajuste de columnas, formato numérico/moneda, panel congelado y autofiltro,
de modo que todas las exportaciones (repuestos, maquinaria, reportes) salgan
consistentes y legibles.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ColumnKind = Literal["text", "integer", "decimal", "currency", "percent", "date"]

#: Formatos de celda por tipo de columna.
NUMBER_FORMATS: dict[ColumnKind, str] = {
    "text": "@",
    "integer": "#,##0",
    "decimal": "#,##0.00",
    "currency": '"$"#,##0.00',
    "percent": "0.0%",
    "date": "dd/mm/yyyy",
}

HEADER_FILL = PatternFill("solid", fgColor="4338CA")
TOTAL_FILL = PatternFill("solid", fgColor="EEF2FF")
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Calibri", size=9, color="64748B")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="0F172A")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="4338CA")

_HAIRLINE = Side(style="thin", color="E2E8F0")
CELL_BORDER = Border(left=_HAIRLINE, right=_HAIRLINE, top=_HAIRLINE, bottom=_HAIRLINE)

#: Límites del auto-ajuste, para que una celda larga no genere una columna ilegible.
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 52


@dataclass(frozen=True)
class Column:
    """Definición de una columna de exportación."""

    header: str
    kind: ColumnKind = "text"
    width: int | None = None

    @property
    def number_format(self) -> str:
        return NUMBER_FORMATS[self.kind]

    @property
    def alignment(self) -> str:
        if self.kind in ("integer", "decimal", "currency", "percent"):
            return "right"
        if self.kind == "date":
            return "center"
        return "left"


class ExcelWorkbook:
    """Libro XLSX con hojas formateadas de forma homogénea."""

    def __init__(self) -> None:
        self._wb = Workbook()
        # Eliminamos la hoja por defecto: cada hoja se crea explícitamente.
        self._wb.remove(self._wb.active)

    def add_sheet(
        self,
        *,
        title: str,
        columns: Sequence[Column],
        rows: Sequence[Sequence[object]],
        sheet_title: str | None = None,
        subtitle: str | None = None,
        total_row: Sequence[object] | None = None,
        freeze_header: bool = True,
        autofilter: bool = True,
    ) -> Worksheet:
        """Agrega una hoja con título, encabezados y datos ya formateados."""
        ws = self._wb.create_sheet(title=_safe_sheet_name(sheet_title or title))
        last_col = len(columns)

        # --- Bloque de título -------------------------------------------------
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(last_col, 1))
        subtitle_text = subtitle or (
            f"SGMM Portal · Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        ws.cell(row=2, column=1, value=subtitle_text).font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(last_col, 1))
        ws.row_dimensions[1].height = 22

        header_row = 4

        # --- Encabezados ------------------------------------------------------
        for index, column in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=index, value=column.header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[header_row].height = 26

        # --- Datos ------------------------------------------------------------
        for offset, row_values in enumerate(rows):
            excel_row = header_row + 1 + offset
            for index, column in enumerate(columns, start=1):
                raw = row_values[index - 1] if index - 1 < len(row_values) else None
                cell = ws.cell(row=excel_row, column=index, value=_coerce(raw, column.kind))
                cell.font = BODY_FONT
                cell.border = CELL_BORDER
                cell.number_format = column.number_format
                cell.alignment = Alignment(
                    horizontal=column.alignment, vertical="center", wrap_text=False
                )
            if offset % 2 == 1:
                for index in range(1, last_col + 1):
                    ws.cell(row=excel_row, column=index).fill = ZEBRA_FILL

        last_data_row = header_row + len(rows)

        # --- Fila de totales --------------------------------------------------
        # Se deja una fila en blanco de separación: además de leerse mejor, esa
        # fila vacía hace que `read_tabular_upload` cierre el bloque de datos, de
        # modo que reimportar una exportación no interprete "TOTAL" como registro.
        if total_row is not None:
            total_excel_row = last_data_row + 2
            for index, column in enumerate(columns, start=1):
                raw = total_row[index - 1] if index - 1 < len(total_row) else None
                cell = ws.cell(
                    row=total_excel_row, column=index, value=_coerce(raw, column.kind)
                )
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = CELL_BORDER
                cell.number_format = column.number_format
                cell.alignment = Alignment(horizontal=column.alignment, vertical="center")
            last_data_row = total_excel_row

        # --- Presentación -----------------------------------------------------
        _autofit_columns(ws, columns, header_row, last_data_row)

        if freeze_header:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        if autofilter and rows:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(last_col)}{header_row + len(rows)}"
            )

        return ws

    def add_template_sheet(
        self,
        *,
        title: str,
        columns: Sequence[Column],
        example_rows: Sequence[Sequence[object]],
        instructions: Sequence[str],
        sheet_title: str | None = None,
    ) -> Worksheet:
        """Hoja de plantilla de importación, con las instrucciones en hoja aparte.

        Las instrucciones NO pueden vivir en la hoja de datos: al reimportar el
        archivo, el lector las tomaría como filas de registro. Van en una segunda
        hoja para que la primera contenga solo encabezados y ejemplos.
        """
        ws = self.add_sheet(
            title=title,
            columns=columns,
            rows=example_rows,
            sheet_title=sheet_title,
            subtitle=(
                "Complete una fila por registro. No modifique los encabezados. "
                "Consulte la hoja 'Instrucciones'."
            ),
        )

        guide = self._wb.create_sheet(title="Instrucciones")
        guide.cell(row=1, column=1, value=f"{title} · Instrucciones").font = TITLE_FONT
        for offset, line in enumerate(instructions, start=1):
            cell = guide.cell(row=2 + offset, column=1, value=f"• {line}")
            cell.font = SUBTITLE_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        guide.column_dimensions["A"].width = 110

        return ws

    def render(self) -> bytes:
        """Serializa el libro a bytes listos para descargar."""
        if not self._wb.sheetnames:
            self._wb.create_sheet(title="Sin datos")
        buffer = io.BytesIO()
        self._wb.save(buffer)
        return buffer.getvalue()


def _coerce(value: object, kind: ColumnKind) -> object:
    """Convierte el valor al tipo nativo que Excel necesita para formatear bien."""
    if value is None or value == "":
        return None
    if kind in ("integer",):
        try:
            return int(float(value))  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
    if kind in ("decimal", "currency"):
        try:
            return round(float(value), 2)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
    if kind == "percent":
        try:
            # Excel interpreta el formato de porcentaje sobre la fracción.
            return float(value) / 100.0  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
    if kind == "date":
        if isinstance(value, (datetime, date)):
            return value
        return str(value)
    return str(value)


def _autofit_columns(
    ws: Worksheet, columns: Sequence[Column], header_row: int, last_row: int
) -> None:
    """Auto-ajusta el ancho de cada columna al contenido más largo, acotado."""
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        if column.width:
            ws.column_dimensions[letter].width = column.width
            continue

        longest = len(column.header)
        for row in range(header_row + 1, last_row + 1):
            value = ws.cell(row=row, column=index).value
            if value is None:
                continue
            if isinstance(value, float):
                rendered = f"{value:,.2f}"
            elif isinstance(value, int):
                rendered = f"{value:,}"
            elif isinstance(value, (datetime, date)):
                rendered = "00/00/0000"
            else:
                rendered = str(value)
            longest = max(longest, len(rendered))

        ws.column_dimensions[letter].width = max(
            MIN_COLUMN_WIDTH, min(longest + 3, MAX_COLUMN_WIDTH)
        )


def _safe_sheet_name(name: str) -> str:
    """Excel limita los nombres de hoja a 31 caracteres y prohíbe algunos símbolos."""
    cleaned = "".join("-" if c in "[]:*?/\\" else c for c in name)
    return cleaned[:31] or "Hoja"


def read_tabular_upload(content: bytes, filename: str) -> list[dict[str, str]]:
    """Lee un archivo de importación (XLSX o CSV) como lista de diccionarios.

    Permite que un mismo endpoint acepte la plantilla Excel que genera el sistema
    y también un CSV equivalente, que es lo que pide la spec 4.4.
    """
    lower = (filename or "").lower()

    if lower.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        header_index = _find_header_row(rows)
        if header_index is None:
            return []

        headers = [
            str(cell).strip() if cell is not None else ""
            for cell in rows[header_index]
        ]
        records: list[dict[str, str]] = []
        for row in rows[header_index + 1 :]:
            # Una fila totalmente vacía cierra el bloque de datos: lo que venga
            # después (notas, totales manuales, comentarios) no son registros.
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                break
            record = {
                header: ("" if row[i] is None else str(row[i]).strip())
                for i, header in enumerate(headers)
                if header and i < len(row)
            }
            records.append(record)
        return records

    # CSV (utf-8-sig tolera el BOM que agrega Excel al guardar).
    import csv

    decoded = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(decoded))
    return [
        {(k or "").strip(): (v or "").strip() for k, v in row.items()}
        for row in reader
    ]


# ---------------------------------------------------------------------------
# Lectura de valores numéricos de un archivo importado
# ---------------------------------------------------------------------------
def parse_decimal_cell(raw: object, *, field: str, default: float = 0.0) -> float:
    """Interpreta una celda numérica de un archivo importado, tolerando formatos.

    ``read_tabular_upload`` entrega todas las celdas como texto, y openpyxl
    devuelve los números de Excel como ``float``: un stock de 5 llega como
    ``"5.0"``. Además, un CSV guardado desde Excel puede traer separador de
    miles (``"1,234.56"``), coma decimal (``"1.234,56"``), símbolo de moneda o
    un negativo entre paréntesis. Se normaliza todo aquí para que ningún
    endpoint de importación tenga que repetir estas reglas.

    Lanza ``ValueError`` con un mensaje en español si el valor no es numérico.
    """
    if raw is None:
        return default
    if isinstance(raw, bool):
        # Un booleano no es una cantidad: se rechaza antes de que int(True) == 1.
        raise ValueError(f"'{field}' debe ser numérico (recibido: '{raw}').")
    if isinstance(raw, (int, float)):
        return float(raw)

    token = str(raw).strip()
    if not token:
        return default

    # Negativos contables: (1.234,56) equivale a -1234,56.
    negative = token.startswith("(") and token.endswith(")")
    if negative:
        token = token[1:-1].strip()

    # Se descarta todo lo que no sea dígito, separador o signo (moneda, unidades,
    # espacios de agrupación incluido el espacio fino que insertan algunas hojas).
    cleaned = "".join(c for c in token if c.isdigit() or c in ",.-+")
    if not cleaned or cleaned in ("-", "+", ".", ","):
        raise ValueError(f"'{field}' debe ser numérico (recibido: '{raw}').")

    cleaned = _normalize_decimal_separator(cleaned)

    try:
        value = float(cleaned)
    except ValueError as exc:
        raise ValueError(
            f"'{field}' debe ser numérico (recibido: '{raw}')."
        ) from exc

    return -value if negative else value


def parse_int_cell(raw: object, *, field: str, default: int = 0) -> int:
    """Igual que :func:`parse_decimal_cell` pero devuelve un entero.

    Un valor con decimales significativos se rechaza en lugar de truncarse en
    silencio: en cantidades de inventario, aceptar "2,5 unidades" como 2
    falsearía el stock sin avisar. En cambio ``"5.0"`` sí es un 5 legítimo, que
    es la forma en la que Excel entrega los enteros.
    """
    value = parse_decimal_cell(raw, field=field, default=float(default))
    rounded = round(value)
    if abs(value - rounded) > 1e-9:
        raise ValueError(
            f"'{field}' debe ser un número entero (recibido: '{raw}')."
        )
    return int(rounded)


def _normalize_decimal_separator(token: str) -> str:
    """Resuelve si la coma o el punto es el separador decimal.

    Reglas, en orden:

    1. Si aparecen AMBOS separadores, manda el último y el otro es de miles.
       Cubre ``1,234.56`` (inglés) y ``1.234,56`` (español).
    2. Si el mismo separador aparece varias veces, solo puede ser de miles
       (``1.234.567``).
    3. Si aparece una sola vez, se interpreta como separador DECIMAL.

    La regla 3 es deliberada: ``"1.234"`` es ambiguo (mil doscientos treinta y
    cuatro, o uno con 234 milésimas). Tomarlo como decimal nunca multiplica el
    valor por mil a espaldas del usuario, y en las cantidades enteras
    (:func:`parse_int_cell`) el resultado fraccionario se rechaza con un mensaje
    explícito en vez de importar un stock equivocado en silencio.
    """
    last_comma = token.rfind(",")
    last_dot = token.rfind(".")

    if last_comma >= 0 and last_dot >= 0:
        if last_comma > last_dot:
            return token.replace(".", "").replace(",", ".")
        return token.replace(",", "")

    separator = "," if last_comma >= 0 else ("." if last_dot >= 0 else "")
    if not separator:
        return token

    if token.count(separator) > 1:
        return token.replace(separator, "")

    return token.replace(separator, ".")


def _find_header_row(rows: Sequence[Sequence[object]]) -> int | None:
    """Localiza la fila de encabezados, saltando el bloque de título de la plantilla.

    Se considera encabezado la primera fila con dos o más celdas de texto no vacías.
    """
    for index, row in enumerate(rows):
        if row is None:
            continue
        filled = [c for c in row if c is not None and str(c).strip() != ""]
        if len(filled) >= 2:
            return index
    return None
